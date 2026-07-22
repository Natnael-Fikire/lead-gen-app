"""
Native Excel & CSV Export Engine
-----------------------------------
Turns a list of Lead objects into:
  - A polished multi-tab .xlsx workbook (Pandas + openpyxl) with:
      * Dark styled header row, white bold text
      * Zebra-striped data rows
      * Auto-fitted column widths
      * Conditional status coloring for MX verification (green=valid, red=invalid, amber=error/unchecked)
  - A plain .csv (Pandas) for lightweight/portable use.
"""
from __future__ import annotations
import io
from typing import List

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Lead

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # dark slate
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ZEBRA_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")   # light gray
VALID_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # light green
INVALID_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
UNCHECKED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # light amber

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COLUMNS = [
    ("company_name", "Company"),
    ("source_url", "Source URL"),
    ("domain", "Domain"),
    ("emails", "Emails"),
    ("phones", "Phones"),
    ("addresses", "Address"),
    ("social_links", "Social Links"),
    ("mx_status", "MX Status"),
]


def _leads_to_dataframe(leads: List[Lead]) -> pd.DataFrame:
    rows = []
    for lead in leads:
        rows.append({
            "company_name": lead.company_name or "",
            "source_url": lead.source_url,
            "domain": lead.domain or "",
            "emails": "; ".join(lead.emails),
            "phones": "; ".join(lead.phones),
            "addresses": "; ".join(lead.addresses),
            "social_links": "; ".join(lead.social_links),
            "mx_status": (lead.mx_status or "unchecked").upper(),
        })
    df = pd.DataFrame(rows, columns=[c[0] for c in COLUMNS])
    df.columns = [c[1] for c in COLUMNS]
    return df


def build_xlsx(leads: List[Lead]) -> bytes:
    df = _leads_to_dataframe(leads)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")

        # Summary tab
        total = len(leads)
        valid = sum(1 for l in leads if (l.mx_status or "").lower() == "valid")
        invalid = sum(1 for l in leads if (l.mx_status or "").lower() == "invalid")
        unchecked = total - valid - invalid
        summary_df = pd.DataFrame({
            "Metric": ["Total Leads", "Valid MX (deliverable domain)", "Invalid MX", "Unchecked/Error"],
            "Value": [total, valid, invalid, unchecked],
        })
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        workbook = writer.book
        ws = writer.sheets["Leads"]
        _style_leads_sheet(ws, df)
        _style_summary_sheet(writer.sheets["Summary"], summary_df)

    buffer.seek(0)
    return buffer.getvalue()


def _style_leads_sheet(ws, df: pd.DataFrame) -> None:
    n_rows, n_cols = df.shape

    # Header row styling
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    mx_col_idx = df.columns.get_loc("MX Status") + 1  # 1-indexed

    # Data rows: zebra stripe + MX status color + borders
    for row_idx in range(2, n_rows + 2):
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx == mx_col_idx:
                status = str(cell.value or "").lower()
                if status == "valid":
                    cell.fill = VALID_FILL
                elif status == "invalid":
                    cell.fill = INVALID_FILL
                else:
                    cell.fill = UNCHECKED_FILL
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_even:
                cell.fill = ZEBRA_FILL

    # Auto-fit column widths based on content length
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, n_rows + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def _style_summary_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx in range(1, df.shape[1] + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for row_idx in range(2, df.shape[0] + 2):
        for col_idx in range(1, df.shape[1] + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(20, len(col_name) + 4)


def build_csv(leads: List[Lead]) -> bytes:
    df = _leads_to_dataframe(leads)
    return df.to_csv(index=False).encode("utf-8")
